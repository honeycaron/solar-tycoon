import time


from openpyxl.workbook import Workbook
from selenium.webdriver.common.by import By
from seleniumwire import webdriver
import pandas as pd
from time import sleep

# 드라이버 실행
options = webdriver.ChromeOptions()
options.add_argument("window-size=1920x1080")
driver = webdriver.Chrome(options=options)
driver.get('https://www.google.co.kr/maps/place/%EC%9A%B0%EC%A7%84%ED%95%B4%EC%9E%A5%EA%B5%AD/data=!4m16!1m9!3m8!1s0x350ce4ab27b511d5:0xb6bd54bc3de91ebb!2z7Jqw7KeE7ZW07J6l6rWt!8m2!3d33.511505!4d126.5200319!9m1!1b1!16s%2Fg%2F1tlqmr7v!3m5!1s0x350ce4ab27b511d5:0xb6bd54bc3de91ebb!8m2!3d33.511505!4d126.5200319!16s%2Fg%2F1tlqmr7v?entry=ttu')

def detail_btn():
    driver.find_element(By.CSS_SELECTOR, '.w8nwRe.kyuRq').click()
    time.sleep(2)


def review_scroll():
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        # 스크롤을 아래로 내림
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        time.sleep(2)

        # 새 높이 계산
        new_height = driver.execute_script("return document.body.scrollHeight")

        review_dates = driver.find_elements(By.XPATH, '//span[contains(@class, "ODSEW-ShBeI-RgZmSc-date")]')
        for date in review_dates:
            review_date_text = date.text
            if '6달 전' in review_date_text:
                print("6달 전 리뷰 발견, 스크롤 종료")
                return  # 함수를 종료하여 스크롤을 멈춤

        # 더 이상 스크롤할 내용이 없을 때 종료
        if new_height == last_height:
            break

        last_height = new_height

def google_reviews(store_list, store_call, jeju, count=10):
    #storeNm = driver.find_element(By.XPATH,'//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[1]/h1').text
    xlsx = Workbook()
    #list_sheet = xlsx.create_sheet(storeNm)
    list_sheet = xlsx.create_sheet("우진해장국")
    list_sheet.append(['nickname', 'date','content'])

    # 리뷰 버튼 클릭
    try:
        review_btn = driver.find_element(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[3]/div/div/button[2]/div[2]/div[2]')
        review_btn.click()
    except Exception as e:
        print(f"Error 리뷰버튼: {e}")


    # 정렬 버튼 클릭 및 최신순 정렬
    try:
        sort_btn = driver.find_element(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[7]/div[2]/button/span/span[2]')
        sort_btn.click()
        recent_btn = driver.find_element(By.CSS_SELECTOR, '#action-menu > div:nth-child(2)')
        recent_btn.click()
    except Exception as e:
        print(f"정렬 Error")

    sleep(2)

    # 리뷰 더보기 펼치기
    review_scroll()


    reviews = driver.find_elements(By.CSS_SELECTOR, '.jftiEf.fontBodyMedium')

    print(len(reviews))
    result_list = []
    for i in reviews:
        nickname = i.find_element(By.CSS_SELECTOR, '.d4r55')
        detail_btn()
        content = i.find_element(By.CLASS_NAME,'MyEned')
        date = i.find_element(By.CSS_SELECTOR,'.rsqaWe')

        # exception handling
        nickname = nickname.text if nickname else ''
        content = content.text if content else ''
        date = date.text if date else ''
        time.sleep(1)

        print(nickname, '/', date, '/', content)
        result_list.append([nickname, date, content])
        time.sleep(1)

    df = pd.DataFrame(result_list)
    #filename = f'{storeNm}.csv'
    filename = '우진해장국.csv'
    df.to_csv(filename, encoding='utf-8-sig', index=False)
    print(f"Saved reviews to {filename}")

    driver.quit()

# 사용 예제
df_restNm = pd.read_excel("제주도_restnum_with_adr_pn.xlsx")
store_list = df_restNm['업체명']
store_call = df_restNm['전화번호']

# 함수 실행
google_reviews(store_list, store_call,'제주도', count=5)
